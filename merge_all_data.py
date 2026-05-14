"""
merge_all_data.py — دمج جميع ملفات Excel في قاعدة البيانات
===========================================================
يستورد:
  1. October Memberships.xlsx  → skaters + payments + bundles
  2. August 2025 Attendence.xlsx + October 2025 Attendence.xlsx → attendance
  3. Off-Ice Attendence.xlsx   → attendance (off-ice sessions)
  4. FSE skills.xlsx           → skating_elements + player_skills
  5. FSE1.xlsx                 → level definitions + membership levels
  6. Data.xlsx                 → bundles (pricing)
  7. برنامج_تدريب_التزلج_على_الجليد.xlsx → skaters (players list)
  8. نظام_التدريب_الشامل_المتقدم_عربي.xlsx → skating_elements (YouTube links)
  9. نظام_التدريب_المتطور_CoachJulia.xlsx  → skating_elements (CoachJulia links)

Usage:
    python merge_all_data.py            # merge everything
    python merge_all_data.py --dry-run  # show what would be imported, no DB writes
"""

import argparse
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

ROOT   = Path(__file__).parent
DB     = ROOT / "skating_database.db"

# ── helpers ────────────────────────────────────────────────────────────────

def get_conn():
    conn = sqlite3.connect(str(DB))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def create_tables(conn):
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS skaters (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        coach_name TEXT,
        level TEXT,
        bundle TEXT,
        membership_status TEXT DEFAULT 'active',
        gender TEXT,
        notes TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS attendance (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        skater_name TEXT NOT NULL,
        skater_id INTEGER,
        date TEXT NOT NULL,
        session_type TEXT DEFAULT 'on-ice',
        status TEXT DEFAULT 'present',
        coach TEXT,
        recorded_by TEXT
    );

    CREATE TABLE IF NOT EXISTS payments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        skater_id INTEGER,
        skater_name TEXT,
        amount REAL,
        payment_date TEXT,
        payment_method TEXT,
        bundle_type TEXT,
        discount REAL DEFAULT 0,
        received_by TEXT,
        status TEXT DEFAULT 'completed',
        created_at TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS bundles (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        bundle_type TEXT,
        hours INTEGER,
        price REAL,
        rink TEXT,
        group_level TEXT,
        coach TEXT,
        is_active INTEGER DEFAULT 1,
        created_at TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS skating_elements (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name_ar TEXT,
        name_en TEXT NOT NULL,
        isu_code TEXT UNIQUE,
        element_type TEXT,
        sub_type TEXT,
        level TEXT,
        base_value REAL,
        difficulty TEXT,
        description TEXT,
        youtube_url TEXT,
        notes TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS player_skills (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        skater_name TEXT NOT NULL,
        skater_id INTEGER,
        element_name TEXT NOT NULL,
        element_id INTEGER,
        fse_level TEXT,
        achieved INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now')),
        UNIQUE(skater_name, element_name)
    );

    CREATE TABLE IF NOT EXISTS training_videos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        filepath TEXT UNIQUE,
        filename TEXT,
        source TEXT DEFAULT 'local',
        youtube_url TEXT,
        label TEXT,
        element_id INTEGER,
        rotations INTEGER,
        duration REAL,
        width INTEGER,
        height INTEGER,
        fps REAL,
        size_mb REAL,
        used_in_training INTEGER DEFAULT 0,
        created_at TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS discovered_videos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        filepath TEXT UNIQUE,
        filename TEXT,
        size_mb REAL,
        duration REAL,
        width INTEGER,
        height INTEGER,
        fps REAL,
        label TEXT DEFAULT NULL,
        analyzed INTEGER DEFAULT 0,
        scan_date TEXT
    );
    """)
    conn.commit()


def wb_rows(path, sheet=0, skip=1):
    """Yield rows as lists from an xlsx sheet, skipping header rows."""
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.worksheets[sheet] if isinstance(sheet, int) else wb[sheet]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < skip:
            continue
        yield list(row)
    wb.close()


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s.lower() not in ('none', 'nan', '-', 'n/a') else None


def normalize_name(n):
    if not n:
        return None
    return ' '.join(str(n).strip().split())


def to_date(v):
    if not v:
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s[:10], fmt[:len(s[:10])]).strftime('%Y-%m-%d')
        except Exception:
            continue
    return s[:10] if len(s) >= 10 else None


# ── 1. October Memberships ─────────────────────────────────────────────────

def import_memberships(conn, dry_run=False):
    path = ROOT / "October Memberships.xlsx"
    if not path.exists():
        print(f"  [skip] {path.name} not found")
        return 0, 0

    skaters_added = 0
    payments_added = 0

    for row in wb_rows(path, 0, skip=1):
        if not any(row):
            continue
        # Member, Level, Coach, Bundle, Amount, Discount, DOP, Paid to, Pain in
        name     = normalize_name(row[0])
        level    = clean(row[1])
        coach    = clean(row[2])
        bundle   = clean(row[3])
        amount   = row[4]
        discount = row[5]
        dop      = to_date(row[6])
        paid_to  = clean(row[7])
        method   = clean(row[8])

        if not name:
            continue

        if not dry_run:
            conn.execute("""
                INSERT OR IGNORE INTO skaters (name, level, coach_name, bundle)
                VALUES (?, ?, ?, ?)
            """, (name, level, coach, bundle))
            conn.execute("""
                UPDATE skaters SET level=COALESCE(?,level),
                  coach_name=COALESCE(?,coach_name), bundle=COALESCE(?,bundle)
                WHERE name=?
            """, (level, coach, bundle, name))
            skaters_added += 1

            if amount:
                conn.execute("""
                    INSERT OR IGNORE INTO payments
                      (skater_name, amount, payment_date, payment_method,
                       bundle_type, discount, received_by)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (name, float(str(amount).replace(',','') or 0),
                      dop, method, bundle,
                      float(str(discount).replace(',','') or 0) if discount else 0,
                      paid_to))
                payments_added += 1

    if not dry_run:
        conn.commit()
    print(f"  Memberships: {skaters_added} skaters, {payments_added} payments")
    return skaters_added, payments_added


# ── 2. Attendance files ────────────────────────────────────────────────────

def import_attendance_wide(conn, path, dry_run=False):
    """
    Wide-format attendance: row 0 = dates, row 1 = Off-ice/On-ice/Coach headers,
    row 2 = first coach row, row 3+ = skater names with Y/present values.
    """
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 3:
        return 0

    # Row 0: dates in columns 1+
    date_row   = rows[0]
    # Row 1: Off-ice / On-ice / Coach repeating
    type_row   = rows[1]

    # Build column map: col_idx → (date_str, session_type)
    col_map = {}
    current_date = None
    for ci, val in enumerate(date_row):
        if ci == 0:
            continue
        if val:
            current_date = to_date(val)
        session_type = clean(type_row[ci]) if ci < len(type_row) else None
        if session_type and session_type.lower() in ('on-ice', 'off-ice', 'off ice', 'on ice'):
            st = 'on-ice' if 'on' in session_type.lower() else 'off-ice'
            col_map[ci] = (current_date, st)

    added = 0
    for row in rows[2:]:
        name = normalize_name(row[0])
        if not name or name.lower() in ('coach', 'trainer', ''):
            continue
        for ci, (date_str, stype) in col_map.items():
            if not date_str:
                continue
            val = clean(row[ci]) if ci < len(row) else None
            if val and val.lower() not in ('0', 'no', 'absent', '', 'none'):
                if not dry_run:
                    conn.execute("""
                        INSERT OR IGNORE INTO attendance
                          (skater_name, date, session_type, status)
                        VALUES (?, ?, ?, 'present')
                    """, (name, date_str, stype))
                    conn.execute("""
                        INSERT OR IGNORE INTO skaters (name) VALUES (?)
                    """, (name,))
                added += 1

    if not dry_run:
        conn.commit()
    return added


def import_all_attendance(conn, dry_run=False):
    files = [
        ROOT / "August 2025 Attendence.xlsx",
        ROOT / "October 2025 Attendence.xlsx",
        ROOT / "Off-Ice Attendence.xlsx",
    ]
    total = 0
    for f in files:
        if f.exists():
            n = import_attendance_wide(conn, f, dry_run)
            print(f"  Attendance {f.name}: {n} records")
            total += n
        else:
            print(f"  [skip] {f.name} not found")
    return total


# ── 3. Data.xlsx → bundles ─────────────────────────────────────────────────

def import_bundles(conn, dry_run=False):
    path = ROOT / "Data.xlsx"
    if not path.exists():
        print(f"  [skip] Data.xlsx not found")
        return 0

    added = 0
    for row in wb_rows(path, 0, skip=1):
        # Bundles, Coach, Rink, Group, Level, Coach
        # Type, Hours, Fees, ELMalahy, Kids, Kids, Ahmed/clara
        name   = clean(row[0])
        hours  = clean(row[1])
        price  = clean(row[2])
        rink   = clean(row[3])
        group  = clean(row[4])
        level  = clean(row[5])
        coach  = clean(row[6])

        if not name or name.lower() in ('bundles', 'type'):
            continue

        if not dry_run:
            conn.execute("""
                INSERT OR IGNORE INTO bundles
                  (name, hours, price, rink, group_level, coach)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (name,
                  int(hours) if hours and str(hours).isdigit() else None,
                  float(str(price).replace(',','')) if price and str(price).replace(',','').replace('.','').isdigit() else None,
                  rink, level or group, coach))
        added += 1

    if not dry_run:
        conn.commit()
    print(f"  Bundles: {added} records")
    return added


# ── 4. FSE skills.xlsx → elements + player_skills ─────────────────────────

def import_fse_skills(conn, dry_run=False):
    path = ROOT / "FSE skills.xlsx"
    if not path.exists():
        print(f"  [skip] FSE skills.xlsx not found")
        return 0, 0

    wb   = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws   = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return 0, 0

    # Row 0: TESTING LEVEL, Pre-Alpha, Alpha, Beta, Gamma, Delta, FS1…FS10
    level_row  = rows[0]
    # Row 1: SKILLS, skill1, skill2 …
    skills_row = rows[1]

    # Build column → (fse_level, skill_name) mapping
    col_map = {}
    for ci in range(1, len(skills_row)):
        skill = clean(skills_row[ci])
        level = clean(level_row[ci]) if ci < len(level_row) else None
        if skill:
            col_map[ci] = (level or 'Unknown', skill)

    # Insert elements
    elements_added = 0
    for fse_level, skill_name in col_map.values():
        if not dry_run:
            conn.execute("""
                INSERT OR IGNORE INTO skating_elements
                  (name_en, element_type, level)
                VALUES (?, 'skill', ?)
            """, (skill_name, fse_level))
        elements_added += 1

    # Row 2+: player name in col 0, Y/blank per skill
    skills_added = 0
    for row in rows[2:]:
        name = normalize_name(row[0])
        if not name:
            continue
        for ci, (fse_level, skill_name) in col_map.items():
            val = clean(row[ci]) if ci < len(row) else None
            achieved = 1 if val and val.upper() in ('Y', 'YES', '1', 'TRUE', '✓') else 0
            if not dry_run:
                conn.execute("""
                    INSERT OR IGNORE INTO player_skills
                      (skater_name, element_name, fse_level, achieved)
                    VALUES (?, ?, ?, ?)
                """, (name, skill_name, fse_level, achieved))
                conn.execute("""
                    INSERT OR IGNORE INTO skaters (name) VALUES (?)
                """, (name,))
            skills_added += 1

    if not dry_run:
        conn.commit()
    print(f"  FSE skills: {elements_added} elements, {skills_added} skill records")
    return elements_added, skills_added


# ── 5. Training system Excel → skating_elements with YouTube links ─────────

def import_elements_youtube(conn, dry_run=False):
    sources = [
        (ROOT / "نظام_التدريب_الشامل_المتقدم_عربي.xlsx",
         "🦘 القفزات - Jumps",
         # cols: name_ar, name_en, code, type, level, base_value, youtube, notes
         {'name_ar':0,'name_en':1,'code':2,'sub_type':3,'level':4,
          'base_value':5,'youtube':6,'notes':7},
         'jump'),
        (ROOT / "نظام_التدريب_الشامل_المتقدم_عربي.xlsx",
         "💫 الدورانات - Spins",
         # cols: name_ar, name_en, code, position, level, youtube, tech_points
         {'name_ar':0,'name_en':1,'code':2,'sub_type':3,'level':4,
          'youtube':5,'notes':6},
         'spin'),
        (ROOT / "نظام_التدريب_المتطور_CoachJulia.xlsx",
         "العناصر💪",
         # cols: element, code, description, difficulty, level, youtube
         {'name_en':0,'code':1,'notes':2,'difficulty':3,'level':4,'youtube':5},
         None),   # type inferred from section headers
    ]

    added = 0
    for fpath, sheet_name, col_map, default_type in sources:
        if not fpath.exists():
            print(f"  [skip] {fpath.name} not found")
            continue
        try:
            wb = openpyxl.load_workbook(str(fpath), read_only=True, data_only=True)
            ws = wb[sheet_name]
            current_type = default_type
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue   # header
                row = list(row)

                # Section header detection (CoachJulia sheet)
                name_en = clean(row[col_map.get('name_en', 0)])
                if name_en and name_en.upper() in ('⛸️ JUMPS', 'JUMPS', '🌀 SPINS', 'SPINS',
                                                    '🔄 TURNS', 'TURNS', '🏃 STEPS', 'STEPS'):
                    current_type = ('jump' if 'JUMP' in name_en.upper() else
                                    'spin' if 'SPIN' in name_en.upper() else 'step')
                    continue

                if not name_en:
                    continue

                code     = clean(row[col_map.get('code', 99)]) if col_map.get('code',99) < len(row) else None
                name_ar  = clean(row[col_map.get('name_ar',99)]) if col_map.get('name_ar',99) < len(row) else None
                sub_type = clean(row[col_map.get('sub_type',99)]) if col_map.get('sub_type',99) < len(row) else None
                level    = clean(row[col_map.get('level',99)]) if col_map.get('level',99) < len(row) else None
                bv_raw   = row[col_map.get('base_value',99)] if col_map.get('base_value',99) < len(row) else None
                youtube  = clean(row[col_map.get('youtube',99)]) if col_map.get('youtube',99) < len(row) else None
                diff     = clean(row[col_map.get('difficulty',99)]) if col_map.get('difficulty',99) < len(row) else None
                notes    = clean(row[col_map.get('notes',99)]) if col_map.get('notes',99) < len(row) else None

                try:
                    base_value = float(str(bv_raw)) if bv_raw and str(bv_raw) not in ('-','') else None
                except Exception:
                    base_value = None

                # Skip YouTube channel homepage — not a specific video
                if youtube and youtube.endswith('@CoachJulia'):
                    youtube = 'https://www.youtube.com/@CoachJulia'

                if not dry_run:
                    conn.execute("""
                        INSERT OR IGNORE INTO skating_elements
                          (name_ar, name_en, isu_code, element_type, sub_type,
                           level, base_value, difficulty, youtube_url, notes)
                        VALUES (?,?,?,?,?,?,?,?,?,?)
                    """, (name_ar, name_en, code, current_type, sub_type,
                          level, base_value, diff, youtube, notes))
                    # Update youtube if already exists
                    if youtube:
                        conn.execute("""
                            UPDATE skating_elements
                            SET youtube_url = COALESCE(youtube_url, ?),
                                name_ar     = COALESCE(name_ar, ?),
                                base_value  = COALESCE(base_value, ?)
                            WHERE name_en = ?
                        """, (youtube, name_ar, base_value, name_en))
                added += 1
            wb.close()
        except Exception as e:
            print(f"  WARNING {fpath.name}/{sheet_name}: {e}")

    if not dry_run:
        conn.commit()
    print(f"  Skating elements: {added} records (with YouTube links)")
    return added


# ── 6. Arabic training plan → skaters ─────────────────────────────────────

def import_arabic_training(conn, dry_run=False):
    path = ROOT / "برنامج_تدريب_التزلج_على_الجليد.xlsx"
    if not path.exists():
        print(f"  [skip] {path.name} not found")
        return 0

    added = 0
    for row in wb_rows(path, 'قائمة اللاعبين', skip=1):
        # الاسم, المستوى, مستوى التدريب, المدرب, حصص على الجليد, حصص خارج الجليد
        name  = normalize_name(row[0])
        level = clean(row[1])
        coach = clean(row[3])
        if not name:
            continue
        if not dry_run:
            conn.execute("""
                INSERT OR IGNORE INTO skaters (name, level, coach_name) VALUES (?,?,?)
            """, (name, level, coach))
            conn.execute("""
                UPDATE skaters SET level=COALESCE(?,level), coach_name=COALESCE(?,coach_name)
                WHERE name=?
            """, (level, coach, name))
        added += 1

    if not dry_run:
        conn.commit()
    print(f"  Arabic training plan: {added} skaters")
    return added


# ── Main ───────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Merge all Excel data into skating_database.db")
    parser.add_argument('--dry-run', action='store_true', help='Show counts without writing to DB')
    args = parser.parse_args()

    print(f"\n{'[DRY RUN] ' if args.dry_run else ''}Merging data into {DB}\n")

    conn = get_conn()
    create_tables(conn)

    print("── Memberships ──────────────────────────────")
    import_memberships(conn, args.dry_run)

    print("\n── Attendance ───────────────────────────────")
    import_all_attendance(conn, args.dry_run)

    print("\n── Bundles / Pricing ────────────────────────")
    import_bundles(conn, args.dry_run)

    print("\n── FSE Skills Matrix ────────────────────────")
    import_fse_skills(conn, args.dry_run)

    print("\n── Skating Elements + YouTube ───────────────")
    import_elements_youtube(conn, args.dry_run)

    print("\n── Arabic Training Plans ────────────────────")
    import_arabic_training(conn, args.dry_run)

    if not args.dry_run:
        # Summary
        tables = ['skaters','attendance','payments','bundles',
                  'skating_elements','player_skills','training_videos']
        print("\n── Database Summary ─────────────────────────")
        for t in tables:
            try:
                n = conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
                print(f"  {t:<22} {n:>6} rows")
            except Exception:
                pass

    conn.close()
    print("\nDone.")


if __name__ == '__main__':
    main()
